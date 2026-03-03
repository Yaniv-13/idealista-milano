#!/usr/bin/env python3
"""
Idealista Milan Rental Search
──────────────────────────────
Scrapes Idealista based on config.yaml, calculates walking distance
to target metro stations, and outputs a dated Excel file.

Run:   python idealista_search.py
       python idealista_search.py --config path/to/other_config.yaml
"""

import argparse
import json
import math
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

import yaml
from bs4 import BeautifulSoup
from commute import get_all_commutes

# Selenium for real-browser fetching (bypasses Cloudflare)
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.by import By
    HAS_SELENIUM = True
except ImportError:
    HAS_SELENIUM = False

try:
    from webdriver_manager.chrome import ChromeDriverManager
    HAS_WDM = True
except ImportError:
    HAS_WDM = False

# requests as fallback only (blocked by Cloudflare on Idealista)
try:
    import requests as _requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Optional: rich console output ─────────────────────────────────────────────
try:
    from rich.console import Console
    from rich.progress import track
    console = Console()
    def log(msg, style=""): console.print(msg, style=style)
except ImportError:
    def log(msg, style=""): print(msg)
    def track(it, description=""): return it

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ══════════════════════════════════════════════════════════════════════════════

def load_config(path: str) -> dict:
    with open(path) as f:
        return yaml.safe_load(f)

# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPING HELPERS
# ══════════════════════════════════════════════════════════════════════════════

# Global browser instance (created once, reused across all fetches)
_BROWSER = None

def get_browser():
    """Return (or create) a headless Chrome instance."""
    global _BROWSER
    if _BROWSER is not None:
        return _BROWSER
    if not HAS_SELENIUM:
        log("❌ selenium not installed. Run: pip3 install selenium", style="red")
        log("   Then install ChromeDriver: https://chromedriver.chromium.org/downloads", style="red")
        sys.exit(1)
    opts = Options()
    # NOTE: NOT headless — Cloudflare blocks headless Chrome.
    # A Chrome window will briefly appear and close when the script finishes.
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1280,900")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
    try:
        if HAS_WDM:
            # webdriver-manager auto-downloads the exact matching ChromeDriver
            service = Service(ChromeDriverManager().install())
            _BROWSER = webdriver.Chrome(service=service, options=opts)
        else:
            _BROWSER = webdriver.Chrome(options=opts)
        _BROWSER.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"}
        )
    except Exception as e:
        log(f"❌ Could not start Chrome: {e}", style="red")
        log("   Fix: pip3 install webdriver-manager", style="yellow")
        sys.exit(1)
    return _BROWSER

def quit_browser():
    global _BROWSER
    if _BROWSER:
        try: _BROWSER.quit()
        except: pass
        _BROWSER = None


def build_filter_slug(cfg: dict) -> str:
    """Build the Idealista filter slug from config."""
    f = cfg["filters"]
    parts = []

    # Price
    if f.get("max_price"):
        parts.append(f"prezzo_{f['max_price']}")

    # Rooms: map to idealista slugs
    min_rooms = f.get("min_rooms", 0)
    room_parts = []
    room_map = {4: "quadrilocali-4", 5: "5-locali-o-piu"}
    if min_rooms <= 4:
        room_parts.append("quadrilocali-4")
        room_parts.append("5-locali-o-piu")
    elif min_rooms == 5:
        room_parts.append("5-locali-o-piu")
    if room_parts:
        parts.extend(room_parts)

    # Bathrooms
    min_baths = f.get("min_bathrooms", 0)
    if min_baths >= 3:
        parts.append("bagno-3")
    elif min_baths == 2:
        parts.append("bagno-2")

    # Pets
    if f.get("pets_allowed"):
        parts.append("animali")

    return ",".join(parts)


def build_urls(cfg: dict) -> list:
    """Build list of URLs to scrape (one per neighborhood, or city-wide).

    Supports single-level slugs  ("porta-vittoria")
    and two-level slugs          ("porta-vittoria/porta-romana").
    Both work as-is since they're inserted directly into the URL path.
    """
    s = cfg["search"]
    lang_prefix = f"/{s['language']}" if s.get("language") else ""
    base = f"https://www.idealista.it{lang_prefix}/{s['operation']}"
    slug = build_filter_slug(cfg)
    filter_path = f"/con-{slug}/" if slug else "/"

    neighborhoods = cfg["filters"].get("neighborhoods", [])
    if neighborhoods:
        return [
            {
                # slug may be "zone" or "zone/subzone" — strip stray slashes, both work
                "url": f"{base}/milano/{n['slug'].strip('/')}{filter_path}",
                "neighborhood": n["label"],
            }
            for n in neighborhoods
        ]
    else:
        city = s["city_slug"]
        return [{"url": f"{base}/{city}{filter_path}", "neighborhood": "All Milan"}]


def fetch_page(url: str) -> Optional[BeautifulSoup]:
    """Fetch a page using a real Chrome browser (bypasses Cloudflare).
    Returns BeautifulSoup, the string "CAPTCHA", or None on error.
    """
    try:
        browser = get_browser()
        browser.get(url)
        time.sleep(4)  # let JS render (Cloudflare needs a moment)

        html = browser.page_source

        # CAPTCHA / Cloudflare block detection
        captcha_signals = [
            "fai scorrere" in html.lower(),
            "slide to verify" in html.lower(),
            "cf-challenge" in html.lower(),
            "just a moment" in html.lower(),
            len(html) < 5000,  # page too short = blocked/empty
        ]
        if any(captcha_signals):
            return "CAPTCHA"

        soup = BeautifulSoup(html, "html.parser")
        return soup
    except Exception as e:
        log(f"  ⚠ Fetch error: {e}", style="yellow")
        return None


def _extract_bathrooms(text: str) -> str:
    if not text:
        return ""
    m = re.search(r"(\d+)\s*(?:bagni|bagno|bathrooms?|baths?)", text, re.I)
    return m.group(1) if m else ""


def _extract_pets_allowed(text: str):
    if not text:
        return None
    t = re.sub(r"\s+", " ", text).strip().lower()
    # Explicit negatives
    if re.search(r"(no\s+pets|pets?\s+not\s+allowed|animali\s+non\s+ammessi|animali\s+vietati)", t):
        return False
    if re.search(r"animali\s+ammessi\s*[:\-]?\s*no", t):
        return False
    # Explicit positives
    if re.search(r"(pets?\s+allowed|pet\s+friendly|animali\s+ammessi|animali\s+consentiti)", t):
        return True
    return None


def _extract_pets_allowed_structured(soup: BeautifulSoup):
    if not soup:
        return None
    # Look for explicit "Animali ammessi" label/value pairs
    for el in soup.find_all(string=re.compile(r"animali", re.I)):
        if not el or not el.strip():
            continue
        parent = el.parent
        full = parent.get_text(" ", strip=True) if parent else el.strip()
        if re.search(r"animali", full, re.I):
            if re.search(r"\b(no|non)\b", full, re.I):
                return False
            if re.search(r"\b(si|yes|ammessi|consentiti)\b", full, re.I):
                return True
        if parent:
            sib = parent.find_next_sibling()
            if sib:
                val = sib.get_text(" ", strip=True)
                if re.search(r"\bno\b", val, re.I):
                    return False
                if re.search(r"\b(si|yes)\b", val, re.I):
                    return True
    return None


def parse_listing_page_details(soup: BeautifulSoup) -> tuple:
    if not soup:
        return "", None
    # Prefer structured sections likely to contain features
    candidates = []
    for sel in [
        ".details__list", ".details", ".feature", ".features",
        ".item-detail", ".item-detail-char", ".item-detail-text",
        "section", "ul", "li"
    ]:
        candidates.extend(soup.select(sel))
    text = " ".join([c.get_text(" ", strip=True) for c in candidates if c]) or soup.get_text(" ", strip=True)
    bathrooms = _extract_bathrooms(text)
    pets_allowed = _extract_pets_allowed_structured(soup)
    if pets_allowed is None:
        pets_allowed = _extract_pets_allowed(text)
    return bathrooms, pets_allowed


def parse_listings(soup: BeautifulSoup, neighborhood: str) -> list:
    """Extract listing data from a search results page."""
    cards = soup.select("article.item")
    listings = []
    for card in cards:
        title_el = card.select_one(".item-link")
        price_el = card.select_one(".item-price")
        detail_el = card.select_one(".item-detail-char")
        desc_el = card.select_one(".item-description p")

        if not title_el:
            continue

        title = title_el.get_text(strip=True)
        # title attribute contains full address e.g. "4 room flat in Via Argelati, 30, Navigli, Milano"
        address_raw = title_el.get("title", "")
        url = title_el.get("href", "")
        if url and not url.startswith("http"):
            url = "https://www.idealista.it" + url

        price_raw = price_el.get_text(strip=True) if price_el else ""
        price_num = int(re.sub(r"[^\d]", "", price_raw)) if price_raw else 0

        detail_parts = []
        if detail_el:
            detail_parts.append(detail_el.get_text(" ", strip=True))
        detail_parts.extend([d.get_text(" ", strip=True) for d in card.select(".item-detail, .item-detail-char, .item-detail-text")])
        detail = " ".join([p for p in detail_parts if p]).strip()
        desc = desc_el.get_text(strip=True) if desc_el else ""

        bathrooms = _extract_bathrooms(detail) or _extract_bathrooms(desc)
        pets_allowed = _extract_pets_allowed(detail)
        if pets_allowed is None:
            pets_allowed = _extract_pets_allowed(desc)

        # Extract rooms from detail only (e.g. "5 rooms
        # Extract rooms from detail only - NOT from title
        rooms_m = re.search(r"^(\d+)\s*room", detail, re.I) or re.search(r"^(\d+)\s*local", detail, re.I)
        sqm_m = re.search(r"(\d+)\s*m", detail)
        floor_m = re.search(r"(\d+)\w*\s*(?:floor|piano)", detail, re.I)

        listing_id = re.search(r"/immobile/(\d+)/", url)

        # Extract neighborhood from address title:
        # e.g. "Flat in Via Argelati, 30, Navigli - Porta Genova, Milano"
        # → "Navigli - Porta Genova"
        nbhd = neighborhood  # fallback
        if address_raw:
            parts = [p.strip() for p in address_raw.split(",")]
            # Neighbourhood is the second-to-last part (before "Milano")
            if len(parts) >= 2:
                candidate = parts[-2].strip()
                # Skip if it looks like a street number or "Milano"
                if candidate and not re.match(r'^[\d\s/]+$', candidate) and candidate.lower() != "milano":
                    nbhd = candidate

        listings.append({
            "id": listing_id.group(1) if listing_id else url,
            "neighborhood": nbhd,
            "title": title,
            "url": url,
            "price_eur": price_num,
            "price_raw": price_raw,
            "rooms": rooms_m.group(1) if rooms_m else "",
            "sqm": sqm_m.group(1) if sqm_m else "",
            "bathrooms": bathrooms,
            "floor": floor_m.group(0) if floor_m else "",
            "has_lift": "Yes" if "lift" in detail.lower() or "ascensore" in detail.lower() else "No",
            "pets_allowed": pets_allowed,
            "desc": desc[:300],
            "address": address_raw,
            "lat": None,
            "lng": None,
        })
    return listings


def get_listing_coordinates(url: str) -> tuple:
    """Extract address from listing page, then geocode via Nominatim (OpenStreetMap).
    Returns (lat, lng) or (None, None).
    """
    try:
        soup = fetch_page(url)
        if not soup or soup == "CAPTCHA":
            return None, None

        # Extract address from page title or heading
        # Title format: "4 room flat for rent in Via Argelati, 30, Navigli, Milan"
        address = None
        title_el = soup.find("title")
        if title_el:
            t = title_el.text
            m = re.search(r'in (.+?),\s*Milan', t, re.I)
            if m:
                address = m.group(1).strip() + ", Milan, Italy"

        if not address:
            h1 = soup.select_one("h1, .main-info__title-minor")
            if h1:
                address = h1.get_text(strip=True) + ", Milan, Italy"

        if not address:
            return None, None

        return geocode_address(address)
    except Exception:
        return None, None


def clean_address(raw: str) -> str:
    """Strip property type prefix and neighbourhood suffix, leaving just street + number."""
    addr = re.sub(r'^.+?\bin\b\s*', '', raw, flags=re.I).strip()
    parts = [p.strip() for p in addr.split(',')][:-1]  # drop last "Milano"
    street_parts = []
    for p in parts:
        is_street  = bool(re.match(r'^(Via|Corso|Piazza|Viale|Largo|Vicolo|Strada|Galleria|Alzaia)', p, re.I))
        is_number  = bool(re.match(r'^[\d\s/a-zA-Z]{1,10}$', p)) and bool(re.search(r'\d', p))
        is_nn      = p.strip().upper() == 'NN'
        is_neighbourhood = not is_street and not is_number and not is_nn and len(street_parts) > 0
        if is_nn:            continue
        if is_neighbourhood: break
        if is_street or is_number or not street_parts:
            street_parts.append(p)
    return ', '.join(street_parts) + ', Milano, Italy' if street_parts else addr

def geocode_address(address: str) -> tuple:
    """Geocode an address using Nominatim (free OpenStreetMap geocoder).
    Cleans the raw Idealista address string before querying.
    Returns (lat, lng) or (None, None).
    """
    if not HAS_REQUESTS:
        return None, None
    try:
        clean = clean_address(address)
        r = _requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": clean, "format": "json", "limit": 1,
                    "countrycodes": "it", "bounded": 1,
                    "viewbox": "9.04,45.39,9.28,45.54"},  # Milan bounding box
            headers={"User-Agent": "Idealista-Milan-Search/1.0"},
            timeout=10,
        )
        results = r.json()
        if results:
            return float(results[0]["lat"]), float(results[0]["lon"])
        return None, None
    except Exception:
        return None, None


# ══════════════════════════════════════════════════════════════════════════════
#  DISTANCE CALCULATION
# ══════════════════════════════════════════════════════════════════════════════

def haversine_meters(lat1, lng1, lat2, lng2) -> float:
    """Straight-line distance between two coords in meters."""
    R = 6_371_000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lng2 - lng1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
    return 2 * R * math.asin(math.sqrt(a))


def walking_distance_ors(lat1, lng1, lat2, lng2, api_key: str) -> Optional[float]:
    """Get walking distance in meters via OpenRouteService API."""
    if not HAS_REQUESTS:
        return None
    try:
        url = "https://api.openrouteservice.org/v2/directions/foot-walking"
        r = _requests.get(url, params={
            "api_key": api_key,
            "start": f"{lng1},{lat1}",
            "end": f"{lng2},{lat2}",
        }, timeout=10)
        data = r.json()
        return data["features"][0]["properties"]["segments"][0]["distance"]
    except Exception:
        return None


def best_metro_distance(listing: dict, stations: list, api_key: str) -> dict:
    """Return the closest metro station and walking distance for a listing."""
    lat, lng = listing.get("lat"), listing.get("lng")
    if lat is None or lng is None:
        return {"closest_station": "N/A", "distance_m": None, "distance_label": "No coords"}

    best_station = None
    best_dist = float("inf")

    for station in stations:
        if api_key:
            dist = walking_distance_ors(lat, lng, station["lat"], station["lng"], api_key)
            if dist is None:
                dist = haversine_meters(lat, lng, station["lat"], station["lng"]) * 1.3  # rough walking factor
        else:
            dist = haversine_meters(lat, lng, station["lat"], station["lng"]) * 1.3

        if dist < best_dist:
            best_dist = dist
            best_station = station["name"]

    mins = round(best_dist / 80)  # ~80m/min walking pace
    label = f"~{round(best_dist)}m / {mins} min walk"
    return {"closest_station": best_station, "distance_m": round(best_dist), "distance_label": label}


# ══════════════════════════════════════════════════════════════════════════════
#  SEEN LISTINGS TRACKER
# ══════════════════════════════════════════════════════════════════════════════

def load_seen(path: Path) -> set:
    if path.exists():
        with open(path) as f:
            return set(json.load(f))
    return set()


def save_seen(path: Path, ids: set):
    with open(path, "w") as f:
        json.dump(list(ids), f)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

DARK   = "1F4E79"
BLUE   = "2E75B6"
LIGHT  = "EBF3FB"
WHITE  = "FFFFFF"
NEW_BG = "E2EFDA"   # green — new listing
OLD_BG = "FFFFFF"   # white — seen before
WARN   = "FCE4D6"   # red-ish — no coords

def _bdr():
    t = Side(style="thin", color="BFBFBF")
    return Border(left=t, right=t, top=t, bottom=t)


def _cell(ws, row, col, val, bold=False, bg=None, color="000000",
          center=False, wrap=False, sz=9, hyperlink=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", size=sz, bold=bold, color=color)
    c.border = _bdr()
    c.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="top", wrap_text=wrap
    )
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if hyperlink:
        c.hyperlink = hyperlink
        c.font = Font(name="Arial", size=sz, color="0563C1", underline="single")
    return c


def write_excel(listings: list[dict], cfg: dict, seen_ids: set, output_path: Path, destinations: list = None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"

    date_str = datetime.now().strftime("%d %b %Y")
    neighborhoods = ", ".join(n["label"] for n in cfg["filters"].get("neighborhoods", [])) or "All Milan"
    # Title
    ws.merge_cells("A1:A1")  # placeholder, overwritten after headers computed
    c = ws["A1"]
    c.value = f"🏠 Milan Rental Search — {date_str}"
    c.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.fill = PatternFill("solid", start_color=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells("A2:A2")  # placeholder, overwritten after headers computed
    c = ws["A2"]
    commute_cfg = cfg.get("commute", {})
    dests = commute_cfg.get("destinations", [])
    filter_dests = [d for d in dests if d.get("filter", False)]
    commute_info = (f"Commute filter: {', '.join(d['label'] + ' ≤' + str(d['max_commute_minutes']) + 'min' for d in filter_dests)}"
                    if filter_dests else "No commute filter")
    c.value = (f"Filters: {cfg['filters'].get('min_rooms',0)}+ rooms, "
               f"{cfg['filters'].get('min_bathrooms',0)}+ baths, pets OK, "
               f"max €{cfg['filters'].get('max_price','?')}/mo, "
               f"{cfg['filters'].get('min_sqm','?')}m²+  |  "
               f"{commute_info}  |  {len(listings)} listings")
    c.font = Font(name="Arial", size=9, italic=True, color="595959")
    c.fill = PatternFill("solid", start_color="D6E4F0")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.append([])  # row 3

    # Headers — base columns + one per commute destination
    destinations = destinations or []
    commute_labels = [d["label"] for d in destinations]
    cols = (["#", "🆕", "Neighborhood", "Price (€/mo)", "Rooms", "m²",
             "Lift", "Floor", "Title", "Link", "Description"]
            + [f"🚇 {lbl}" for lbl in commute_labels])
    widths = ([4, 4, 24, 14, 7, 7, 6, 10, 45, 10, 50]
              + [32] * len(commute_labels))
    total_cols = len(cols)
    # Fix title/subtitle merge range
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
    for i, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = PatternFill("solid", start_color=BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bdr()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 20

    # Sort: new first, then by distance, then price
    def sort_key(l):
        is_new = 0 if l["id"] not in seen_ids else 1
        # Sort by best (lowest) commute time across filter destinations
        commutes = l.get("commutes", [])
        best_commute = min((c["duration_min"] for c in commutes if c["duration_min"] is not None), default=99999)
        return (is_new, best_commute, l["price_eur"])

    listings_sorted = sorted(listings, key=sort_key)

    for i, l in enumerate(listings_sorted):
        row = 5 + i
        is_new = l["id"] not in seen_ids
        has_coords = l.get("lat") is not None
        bg = NEW_BG if is_new else (WARN if not has_coords else OLD_BG)

        p = l["price_eur"]
        price_color = "375623" if p < 3500 else ("7F6000" if p <= 4500 else "C00000")

        vals = [
            (i+1, {}),
            ("🆕" if is_new else "", {"center": True, "color": "375623", "bold": True}),
            (l["neighborhood"], {}),
            (l["price_eur"], {"bold": True, "color": price_color, "bg": bg,
                               "extra": {"number_format": "#,##0"}}),
            (l["rooms"], {"center": True}),
            (l["sqm"], {"center": True}),
            (l["has_lift"], {"center": True,
                              "color": "375623" if l["has_lift"] == "Yes" else "C00000"}),
            (l["floor"], {}),
            (l["title"], {"wrap": True}),
            ("🔗 Open", {"hyperlink": l["url"]}),
            (l["desc"], {"wrap": True}),
        ]
        # Commute columns — summary on first line, full route below
        for commute in l.get("commutes", []):
            mins = commute.get("duration_min")
            summary = commute.get("summary", "N/A")
            steps = commute.get("route_steps", [])
            route_text = (summary + "\n" + "\n".join(steps)) if steps else summary
            if mins is None:
                vals.append((route_text, {"color": "999999", "wrap": True}))
            elif mins <= 20:
                vals.append((route_text, {"color": "375623", "bold": True, "wrap": True}))
            elif mins <= 35:
                vals.append((route_text, {"color": "7F6000", "wrap": True}))
            else:
                vals.append((route_text, {"color": "C00000", "wrap": True}))
        # Pad missing commute columns
        for _ in range(len(commute_labels) - len(l.get("commutes", []))):
            vals.append(("N/A", {"color": "999999", "center": True}))

        for col, (val, kw) in enumerate(vals, 1):
            extra = kw.pop("extra", {})
            c = _cell(ws, row, col, val, bg=kw.get("bg", bg),
                      bold=kw.get("bold", False), color=kw.get("color", "000000"),
                      center=kw.get("center", False), wrap=kw.get("wrap", False),
                      hyperlink=kw.get("hyperlink"))
            for k, v in extra.items():
                setattr(c, k, v)

        ws.row_dimensions[row].height = max(55, 15 * (1 + sum(len(c.get("route_steps", [])) for c in l.get("commutes", []))))

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(total_cols)}{4 + len(listings)}"

    # Legend sheet
    ws2 = wb.create_sheet("How to Use")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 60

    def info_row(r, k, v, section=False):
        for col, val in [(1, k), (2, v)]:
            c = ws2.cell(row=r, column=col, value=val)
            if section:
                c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
                c.fill = PatternFill("solid", start_color=BLUE)
            else:
                c.font = Font(name="Arial", size=9)
                if r % 2 == 0:
                    c.fill = PatternFill("solid", start_color=LIGHT)

    ws2.cell(1, 1, "How to Use & Configuration").font = Font(name="Arial", bold=True, size=12)
    rows = [
        ("HOW TO RUN", "", True),
        ("Daily command", "python idealista_search.py"),
        ("Custom config", "python idealista_search.py --config my_config.yaml"),
        ("", ""),
        ("COLOUR LEGEND", "", True),
        ("🟢 Green row", "NEW listing (not seen in previous runs)"),
        ("⬜ White row", "Already seen in a previous run"),
        ("🟠 Orange row", "Listing has no coordinates (distance can't be computed)"),
        ("🔴 Red distance", "Distance exceeds your max_walking_meters setting"),
        ("", ""),
        ("CAPTCHA HANDLING", "", True),
        ("If CAPTCHA appears", "The script will pause and print a URL for you to open manually"),
        ("What to do", "Open the URL in your browser, solve the slider CAPTCHA, then press Enter"),
        ("", ""),
        ("HOW TO CHANGE PARAMETERS", "", True),
        ("Edit config.yaml", "All search parameters live in config.yaml next to this script"),
        ("Neighborhoods", "Add/remove entries under filters.neighborhoods (or leave [] for all Milan)"),
        ("Commute filter", "Edit destinations under commute.destinations — set filter: true to exclude"),
        ("Max commute", "Set max_commute_minutes per destination"),
        ("Max walk", "Set max_walk_minutes per destination — flags with ⚠ if exceeded"),
        ("Include bus", "Set include_bus: true/false under commute"),
        ("", ""),
        ("COMMUTE COLUMN COLOURS", "", True),
        ("🟢 Green", "Under 20 minutes"),
        ("🟡 Yellow", "20–35 minutes"),
        ("🔴 Red", "Over 35 minutes"),
        ("⚠ Warning", "A walking leg exceeded your max_walk_minutes"),
    ]
    for i, args in enumerate(rows, 2):
        info_row(i, args[0], args[1], section=len(args) > 2)

    wb.save(output_path)
    log(f"  ✅ Saved: {output_path}", style="green")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Idealista Milan Rental Search")
    parser.add_argument("--config", default="config.yaml", help="Path to config file")
    parser.add_argument("--no-coords", action="store_true",
                        help="Skip fetching individual listing pages (faster, no distances)")
    args = parser.parse_args()

    cfg = load_config(args.config)
    urls = build_urls(cfg)

    commute_cfg = cfg.get("commute", {})
    gmaps_key = commute_cfg.get("google_maps_api_key", "")
    destinations = commute_cfg.get("destinations", [])
    dep_hour = commute_cfg.get("departure_hour", 9)
    include_bus = commute_cfg.get("include_bus", False)
    filter_destinations = [d for d in destinations if d.get("filter", False)]

    out_cfg = cfg.get("output", {})
    out_folder = Path(out_cfg.get("folder", ".")).expanduser()
    out_folder.mkdir(parents=True, exist_ok=True)
    seen_path = out_folder / out_cfg.get("seen_listings_file", "seen_listings.json")

    seen_ids = load_seen(seen_path) if out_cfg.get("track_new_listings", True) else set()

    date_str = datetime.now().strftime("%Y-%m-%d")
    out_file = out_folder / f"idealista_{date_str}.xlsx"

    log(f"\n[bold]Idealista Search — {date_str}[/bold]", style="cyan")
    log(f"Searching {len(urls)} neighborhood(s)...")

    # Pre-warm browser: open idealista home so Cloudflare sees a real session
    log("  Opening browser and warming up Cloudflare session...")
    warm_url = f"https://www.idealista.it/{'en/' if cfg['search'].get('language') == 'en' else ''}"
    browser = get_browser()
    browser.get(warm_url)
    time.sleep(5)

    # Check if CAPTCHA appears even on home page
    html = browser.page_source
    if any(s in html.lower() for s in ["fai scorrere", "slide to verify", "cf-challenge", "just a moment"]):
        log("\n⚠️  Cloudflare CAPTCHA on startup!", style="bold yellow")
        log("  → Solve the CAPTCHA in the browser window, then press ENTER here.")
        input("   Press ENTER once solved: ")

    log("  ✅ Browser ready\n")

    all_listings = []

    for entry in urls:
        url = entry["url"]
        neighborhood = entry["neighborhood"]
        log(f"  📍 {neighborhood}")
        log(f"     {url}")

        soup = fetch_page(url)

        if soup == "CAPTCHA":
            log(f"\n⚠️  CAPTCHA detected!", style="bold yellow")
            log(f"1. Open this URL in your browser:\n   {url}", style="yellow")
            log("2. Solve the slider CAPTCHA (slide the arrow to the right)")
            log("3. Press ENTER here once you've solved it...")
            input()
            soup = fetch_page(url)
            if soup == "CAPTCHA" or not soup:
                log(f"  ✗ Still blocked. Skipping {neighborhood}.", style="red")
                continue

        if not soup:
            log(f"  ✗ Failed to fetch. Skipping.", style="red")
            continue

        listings = parse_listings(soup, neighborhood)

        if len(listings) == 0:
            # Could be Cloudflare silently blocking — prompt user to verify
            log(f"\n⚠️  0 listings found — may be a Cloudflare block.", style="bold yellow")
            log(f"1. Check the browser window that opened — does it show listings?", style="yellow")
            log(f"2. If you see a CAPTCHA or blank page, solve it then press ENTER.")
            log(f"3. If the page looks normal (no listings genuinely), just press ENTER.")
            input("   → Press ENTER to continue: ")
            # Re-fetch after user has handled any CAPTCHA
            soup = fetch_page(url)
            if soup and soup != "CAPTCHA":
                listings = parse_listings(soup, neighborhood)
                log(f"  → {len(listings)} listings after retry")

        log(f"  → {len(listings)} listings found")
        all_listings.extend(listings)
        time.sleep(1.5)  # polite delay

    if not all_listings:
        log("\n❌ No listings found. Check your config or CAPTCHA status.", style="red")
        sys.exit(1)

    log(f"\nTotal: {len(all_listings)} listings across all neighborhoods")

    # ── Min sqm post-filter (Idealista has no URL slug for this) ──────────
    min_sqm = cfg["filters"].get("min_sqm")
    if min_sqm:
        before_sqm = len(all_listings)
        all_listings = [l for l in all_listings if not l["sqm"] or int(l["sqm"]) >= min_sqm]
        dropped = before_sqm - len(all_listings)
        if dropped:
            log(f"  Filtered out {dropped} listing(s) under {min_sqm}m²")

    # ── Geocode + commute ──────────────────────────────────────────────────
    need_coords = (gmaps_key and destinations) or not args.no_coords
    if need_coords:
        log("\nGeocoding listing addresses...")
        for listing in all_listings:
            addr = listing.get("address", "")
            if addr:
                lat, lng = geocode_address(addr)
                listing["lat"] = lat
                listing["lng"] = lng
                time.sleep(1.1)  # Nominatim rate limit: 1 req/sec

    if gmaps_key and destinations:
        bus_note = " (incl. bus)" if include_bus else " (metro + tram only)"
        log(f"\nCalculating commute times to {len(destinations)} destination(s){bus_note}...")
        for listing in all_listings:
            commutes = get_all_commutes(listing, destinations, gmaps_key, dep_hour, include_bus)
            listing["commutes"] = commutes
    else:
        for listing in all_listings:
            listing["commutes"] = []
        if destinations and not gmaps_key:
            log("  ℹ️  Add google_maps_api_key to config.yaml to enable commute filtering", style="yellow")

    # ── Commute filter ─────────────────────────────────────────────────────
    if gmaps_key and filter_destinations:
        before = len(all_listings)
        def passes_commute_filter(listing):
            # Keep listings with no coords (can't filter — include with warning)
            if listing.get("lat") is None:
                return True
            for commute in listing.get("commutes", []):
                dest = next((d for d in destinations if d["label"] == commute["label"]), {})
                if not dest.get("filter", False):
                    continue
                # Fails if passes_filter is explicitly False
                if commute.get("passes_filter") is False:
                    return False
            return True
        all_listings = [l for l in all_listings if passes_commute_filter(l)]
        filtered = before - len(all_listings)
        if filtered:
            log(f"  🚫 Filtered out {filtered} listing(s) exceeding commute limits")

    # ── Write Excel ────────────────────────────────────────────────────────
    log(f"\nWriting Excel → {out_file}")
    write_excel(all_listings, cfg, seen_ids, out_file, destinations=destinations)

    # ── Update seen IDs ────────────────────────────────────────────────────
    if out_cfg.get("track_new_listings", True):
        all_ids = {l["id"] for l in all_listings}
        new_count = len(all_ids - seen_ids)
        save_seen(seen_path, seen_ids | all_ids)
        log(f"\n🆕 {new_count} new listings today (highlighted in green)")

    log(f"\n✅ Done! Open: {out_file}", style="bold green")
    quit_browser()


if __name__ == "__main__":
    main()
